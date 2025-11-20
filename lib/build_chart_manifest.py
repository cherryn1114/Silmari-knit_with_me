# lib/build_chart_manifest.py
# 엑셀(Chart Name.xlsx)에 적어 둔 이름을
# assets/chart_from_excel/ 아래의 차트 이미지와
# "순서대로 1:1" 매칭해서 manifest.json 을 만드는 스크립트

from pathlib import Path
import json

import openpyxl

# --------------------------------------------------------------------
# 경로 설정
# --------------------------------------------------------------------
IMG_ROOT   = Path("assets/chart_from_excel")     # 이미지 폴더 루트
NAMES_XLSX = Path("data/Chart Name.xlsx")        # 이름이 적힌 엑셀
OUT_JSON   = IMG_ROOT / "manifest.json"          # 출력 매니페스트

# 시트 이름 ↔ 이미지 폴더 이름 매핑
SHEET_DIR_MAP = {
    "1코 기호":          "1코_기호",
    "1코 2단 기호":      "1코_2단_기호",
    "2코 교차뜨기":      "2코_교차뜨기",
    "3코 교차뜨기":      "3코_교차뜨기",
    "4코 교차뜨기":      "4코_교차뜨기",
    "5코 교차뜨기":      "5코_교차뜨기",
    "6코 교차뜨기":      "6코_교차뜨기",
    "7코 교차뜨기":      "7코_교차뜨기",
    "8코 교차떠기":      "8코_교차뜨기",   # 오타가 있다면 여기 이름만 실제 시트명에 맞게 수정
    "8코 교차뜨기":      "8코_교차뜨기",
    "10코 교차뜨기":     "10코_교차뜨기",
    "3코 방울뜨기":      "3코_방울뜨기",
    "5코 방울뜨기":      "5코_방울뜨기",
    "교차뜨기 일본식 기호": "교차뜨기_일본식_기호",
    "노트뜨기":          "노트뜨기",
}

def clean_text(v):
    """셀 값을 문자열로 정리"""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v)

def main():
    if not NAMES_XLSX.exists():
        raise SystemExit(f"❌ 엑셀 파일을 찾을 수 없습니다: {NAMES_XLSX}")

    wb = openpyxl.load_workbook(NAMES_XLSX, data_only=True)
    manifest = {}

    print(f"📘 엑셀 파일: {NAMES_XLSX}")

    for sheet_title, folder_name in SHEET_DIR_MAP.items():
        if sheet_title not in wb.sheetnames:
            print(f"⚠ 시트 없음: '{sheet_title}' (건너뜀)")
            continue

        img_dir = IMG_ROOT / folder_name
        if not img_dir.exists():
            print(f"⚠ 이미지 폴더 없음: {img_dir} (건너뜀)")
            continue

        # chart_001.png … 순서대로 정렬
        img_files = sorted(
            [p.name for p in img_dir.glob("chart_*.png")]
        )

        ws = wb[sheet_title]

        # 이미지 개수만큼, 2행부터 순서대로 읽기
        names = []
        for i in range(len(img_files)):
            row_idx = 2 + i          # 2행부터 시작
            key  = clean_text(ws.cell(row=row_idx, column=1).value)
            desc = clean_text(ws.cell(row=row_idx, column=2).value)
            names.append({"abbr": key, "desc": desc})

        # 이미지 ↔ 이름 1:1 매칭 (길이는 무조건 동일)
        items = [
            {
                "file": img_files[i],
                "abbr": names[i]["abbr"],
                "desc": names[i]["desc"],
            }
            for i in range(len(img_files))
        ]

        manifest[sheet_title] = {
            "sheet": sheet_title,
            "img_dir": str(img_dir),
            "count_images": len(img_files),
            "count_names": len(names),
            "count_matched": len(items),
            "items": items,
        }

        print(
            f"\n=== 시트: {sheet_title} ===\n"
            f"✔ 이미지 {len(img_files)}개, 엑셀 이름 {len(names)}개 → "
            f"실제 매칭 {len(items)}개 (폴더: {folder_name})"
        )

    # JSON 저장
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\n📁 매니페스트 저장 완료: {OUT_JSON}")

if __name__ == "__main__":
    main()